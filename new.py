# -*- coding: utf-8 -*-
"""
@author: Carlesso
"""
import glob
import csv
from datetime import datetime
from datetime import timedelta
#import dateutil
import pytz
from openpyxl import Workbook

metadata_file = "./stazioni_good.csv"
metadata_dict = dict()
with open(metadata_file) as a_file:
    # csv header is 
    # code;latitudine;longitudine;altitudineslm;strumentazione;regione;comune;area;provincia;label_tmp;label_good
    metadata_reader = csv.DictReader(a_file, delimiter=";", )
    for metadata_row in metadata_reader:
        metadata_dict[ metadata_row["code"] ] =  metadata_row["label_good"]
        

filename_list = glob.glob("input/lmb[0-9][0-9][0-9].csv")
#filename_list = glob.glob("input/lmb080.csv")
output_file = "output/Temp_MetNet_suborarie.xlsx"

one_hour = timedelta(hours=1)
#italy = pytz.timezone("Europe/Rome")
utc = pytz.timezone("UTC")
#it_timezone = dateutil.tz.gettz("Europe/Rome")
# solar time, aka UTC+1, aka Central European Time
solar = pytz.timezone("CET")

# create an Excel Workbook to hold all stations
# TODO maybe set iso_dates=True 
wb = Workbook(write_only=True)
ws0 = wb.create_sheet("metadata")
ws0.append(["Questo file Excel riporta le temperature di alcune stazioni MeteoNetwork."])
ws0.append(["italy_dt è la marca temporale riportata nei dati forniti da MeteoNetwork, che interpretiamo come riferita all'ora Italiana;"])
ws0.append(["L'orario in vigore in Italia si discosta di 1 o 2 ore dall'ora in UTC, a seconda della stagione."])
ws0.append(["utc_td  è la marca temporale riferita a Tempo Universale Coordinato."])
ws0.append(["solar_dt è invece riferita a UTC+1, quindi indipendente dai periodi in cui vige l'ora legale."])
for filename in filename_list:
    # get station_id from filename
    station_id = filename[6:12]
    # convert it to a pretty label
    station_label = metadata_dict[station_id]
    print(station_id, station_label)
    # create a sheet for the station
    ws = wb.create_sheet(title=station_label)
    # append the column header to sheet
    ws.append(["italy_datetime", "solar_datetime", station_label])
    with open(filename) as file:
        # csv header is 
        # code;datetime;temperature
        reader = csv.DictReader(file, delimiter=";", )
        for row in reader:
            # "2013-06-20 00:30:00"
            italy_datetime = datetime.strptime(row["datetime"], "%Y-%m-%d %H:%M:%S")
            solar_datetime = italy_datetime.astimezone(solar)
            try:
                temperature = round(float(row["temperature"]), ndigits=1)
            except: 
                temperature = -999.9
            # append a record 
            out_row = [ italy_datetime, solar_datetime, temperature ]
            ws.append(out_row)
            #print(out_row)
wb.save(output_file)
        
    